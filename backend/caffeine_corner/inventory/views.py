from django.contrib.admin.views.decorators import staff_member_required
from django.shortcuts import render
# inventory/views.py
from django.db.models import Sum, F, Count, Avg
from django.utils import timezone
import datetime
from online_shop.models import Order, Product, LoyaltyPoint, OrderItem, Notification, ActivityLog
from inventory.models import Inventory, PurchaseOrder, PurchaseOrderItem
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models.functions import TruncDay, TruncMonth, TruncWeek
from rest_framework.authentication import SessionAuthentication
from rest_framework_simplejwt.authentication import JWTAuthentication
import csv
import openpyxl
from django.http import HttpResponse
from django.shortcuts import redirect

@staff_member_required
def auto_generate_purchase_orders(request):
    if request.method != 'POST':
        return redirect('/admin/inventory/purchaseorder/')

    # Hanapin lahat ng low stock items na may supplier
    low_stock_items = Inventory.objects.filter(
        quantity_on_hand__lte=F('reorder_points'),
        quantity_on_hand__gt=0,  # hindi pa out of stock
        supplier__isnull=False,  # may supplier
    ).select_related('supplier')

    out_of_stock_items = Inventory.objects.filter(
        quantity_on_hand=0,
        supplier__isnull=False,
    ).select_related('supplier')

    all_items = list(low_stock_items) + list(out_of_stock_items)

    if not all_items:
        from django.contrib import messages
        messages.warning(request, 'No low stock items found that need reordering.')
        return redirect('/admin/inventory/purchaseorder/')

    # Group by supplier
    supplier_items = {}
    for item in all_items:
        supplier_id = item.supplier.id
        if supplier_id not in supplier_items:
            supplier_items[supplier_id] = {
                'supplier': item.supplier,
                'items':    []
            }
        supplier_items[supplier_id]['items'].append(item)

    # Generate PO per supplier
    created_pos = []
    for supplier_id, group in supplier_items.items():
        # Generate PO reference
        today     = timezone.now()
        ref_count = PurchaseOrder.objects.filter(
            ordered_at__year=today.year,
            ordered_at__month=today.month,
        ).count() + 1
        reference = f'PO-{today.strftime("%Y%m")}-{str(ref_count).zfill(4)}'

        # Avoid duplicate PO — check if may existing draft PO for same supplier today
        existing = PurchaseOrder.objects.filter(
            supplier=group['supplier'],
            status='draft',
            ordered_at__date=today.date(),
        ).first()

        if existing:
            po = existing
        else:
            po = PurchaseOrder.objects.create(
                supplier=group['supplier'],
                reference=reference,
                status='draft',
                expected_at=today.date() + timezone.timedelta(days=7),
                created_by=request.user,
                notes=f'Auto-generated on {today.strftime("%B %d, %Y")} for low stock items.',
            )

        # Add items to PO
        for item in group['items']:
            # Skip if already in this PO
            if PurchaseOrderItem.objects.filter(purchase_order=po, inventory=item).exists():
                continue

            PurchaseOrderItem.objects.create(
                purchase_order=po,
                inventory=item,
                quantity_ordered=item.reorder_quantity or (item.reorder_points * 2),
                quantity_received=0,
                unit_cost=item.cost_per_unit,
            )

        created_pos.append(po)

    from django.contrib import messages
    messages.success(
        request,
        f'Successfully generated {len(created_pos)} Purchase Order(s) for {len(all_items)} low stock item(s).'
    )
    return redirect('/admin/inventory/purchaseorder/')

@staff_member_required
def mark_notification_read(request, notification_id):
    from django.shortcuts import redirect
    Notification.objects.filter(id=notification_id).update(is_read=True)
    return redirect('/admin/online_shop/notification/')

class SalesReportView(APIView):
    authentication_classes = [SessionAuthentication, JWTAuthentication]
    permission_classes = [IsAuthenticated]

    def get(self, request):
        period = request.query_params.get('period', 'monthly')
        year   = int(request.query_params.get('year', timezone.now().year))

        orders = Order.objects.filter(
            created_at__year=year,
        )

        total_revenue = orders.aggregate(
            total=Sum(F('items__price') * F('items__quantity'))
        )['total'] or 0
        total_orders  = orders.count()
        total_items   = orders.aggregate(total=Sum('items__quantity'))['total'] or 0

        if period == 'daily':
            trunc = TruncDay
        elif period == 'weekly':
            trunc = TruncWeek
        else:
            trunc = TruncMonth

        revenue_by_period = (
            orders
            .annotate(period=trunc('created_at'))
            .values('period')
            .annotate(
                revenue=Sum(F('items__price') * F('items__quantity')),
                count=Count('id', distinct=True),
            )
            .order_by('period')
        )

        top_products = (
            OrderItem.objects
            .filter(order__created_at__year=year)
            .values('product__id', 'product__name', 'product__image', 'product__category__name')
            .annotate(
                total_sold=Sum('quantity'),
                total_revenue=Sum(F('price') * F('quantity')),
                avg_price=Sum(F('price') * F('quantity')) / Sum('quantity'),
                order_count=Count('order', distinct=True),
            )
            .order_by('-total_revenue')[:10]
        )

        top_categories = (
            OrderItem.objects
            .filter(order__created_at__year=year)
            .values('product__category__name')
            .annotate(
                total_sold=Sum('quantity'),
                total_revenue=Sum(F('price') * F('quantity')),
                product_count=Count('product', distinct=True),
            )
            .order_by('-total_revenue')
        )

        orders_by_status = (
            Order.objects
            .filter(created_at__year=year)
            .values('status')
            .annotate(count=Count('id'))
        )

        # ← Dagdag: Product performance metrics
        product_performance = (
            OrderItem.objects
            .filter(order__created_at__year=year)
            .values(
                'product__id',
                'product__name',
                'product__category__name',
                'product__price',
            )
            .annotate(
                total_sold=Sum('quantity'),
                total_revenue=Sum(F('price') * F('quantity')),
                order_count=Count('order', distinct=True),
                avg_rating=Avg('product__ratings__rating'),
            )
            .order_by('-total_sold')
        )

        return Response({
            'summary': {
                'total_revenue':   total_revenue,
                'total_orders':    total_orders,
                'total_items':     total_items,
                'avg_order_value': round(float(total_revenue) / total_orders, 2) if total_orders else 0,
            },
            'revenue_by_period': [
                {
                    'period':  item['period'].strftime('%b %d' if period == 'daily' else '%b %Y'),
                    'revenue': float(item['revenue'] or 0),
                    'count':   item['count'],
                }
                for item in revenue_by_period
            ],
            'top_products':        list(top_products),
            'top_categories':      list(top_categories),
            'orders_by_status':    list(orders_by_status),
            'product_performance': [
                {
                    'id':           p['product__id'],
                    'name':         p['product__name'],
                    'category':     p['product__category__name'],
                    'base_price':   float(p['product__price'] or 0),
                    'total_sold':   p['total_sold'] or 0,
                    'total_revenue':float(p['total_revenue'] or 0),
                    'order_count':  p['order_count'] or 0,
                    'avg_rating':   round(float(p['avg_rating']), 1) if p['avg_rating'] else None,
                }
                for p in product_performance
            ],
        })


@staff_member_required
def export_orders(request):
    export_format = request.GET.get('format', 'csv')
    status_filter = request.GET.get('status', '')
    date_from     = request.GET.get('date_from', '')
    date_to       = request.GET.get('date_to', '')

    orders = Order.objects.prefetch_related('items__product').order_by('-created_at')

    if status_filter:
        orders = orders.filter(status=status_filter)
    if date_from:
        orders = orders.filter(created_at__date__gte=date_from)
    if date_to:
        orders = orders.filter(created_at__date__lte=date_to)

    if export_format == 'excel':
        return _export_excel(orders)
    return _export_csv(orders)


def _export_csv(orders):
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="orders_{timezone.now().strftime("%Y%m%d")}.csv"'

    writer = csv.writer(response)
    writer.writerow([
        'Order ID', 'Email', 'Order Type', 'Status',
        'Payment Method', 'Payment Status',
        'Products', 'Total Price', 'Discount',
        'Points Earned', 'Address', 'Notes',
        'Event Date', 'Pax', 'Created At'
    ])

    for order in orders:
        products = ', '.join([
            f"{item.product.name} x{item.quantity}"
            for item in order.items.all()
        ])
        writer.writerow([
            f'CC-{str(order.id).zfill(5)}',
            order.email,
            order.get_order_type_display(),
            order.get_status_display(),
            order.get_payment_method_display(),
            order.get_payment_status_display(),
            products,
            order.total_price,
            order.discount,
            order.points_earned,
            order.address,
            order.notes,
            order.event_date or '',
            order.pax or '',
            order.created_at.strftime('%Y-%m-%d %H:%M'),
        ])

    return response


def _export_excel(orders):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Orders'

    header_fill = PatternFill(start_color='3D1F00', end_color='3D1F00', fill_type='solid')
    header_font = Font(color='C4A882', bold=True, size=11)
    border = Border(
        left=Side(style='thin', color='E5E7EB'),
        right=Side(style='thin', color='E5E7EB'),
        top=Side(style='thin', color='E5E7EB'),
        bottom=Side(style='thin', color='E5E7EB'),
    )

    headers = [
        'Order ID', 'Email', 'Order Type', 'Status',
        'Payment Method', 'Payment Status',
        'Products', 'Total Price', 'Discount',
        'Points Earned', 'Address', 'Notes',
        'Event Date', 'Pax', 'Created At'
    ]

    for col, header in enumerate(headers, 1):
        cell           = ws.cell(row=1, column=col, value=header)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = border

    col_widths = [12, 30, 15, 12, 18, 16, 40, 14, 12, 14, 35, 25, 12, 8, 18]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    ws.row_dimensions[1].height = 25

    for row_idx, order in enumerate(orders, 2):
        products = ', '.join([
            f"{item.product.name} x{item.quantity}"
            for item in order.items.all()
        ])
        row_data = [
            f'CC-{str(order.id).zfill(5)}',
            order.email,
            order.get_order_type_display(),
            order.get_status_display(),
            order.get_payment_method_display(),
            order.get_payment_status_display(),
            products,
            float(order.total_price),
            float(order.discount),
            order.points_earned,
            order.address,
            order.notes,
            str(order.event_date) if order.event_date else '',
            order.pax or '',
            order.created_at.strftime('%Y-%m-%d %H:%M'),
        ]

        fill_color = 'FFFFFF' if row_idx % 2 == 0 else 'FAF6F0'
        row_fill   = PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

        for col_idx, value in enumerate(row_data, 1):
            cell           = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = row_fill
            cell.border    = border
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            ws.row_dimensions[row_idx].height = 20

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="orders_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    wb.save(response)
    return response

@staff_member_required
def sales_report_view(request):
    current_year = timezone.now().year
    years = list(range(current_year - 3, current_year + 1))
    return render(request, 'admin/sales_report.html', {
        'current_year': current_year,
        'years': years,
    })

def dashboard_callback(request, context):

    # Stats
    total_orders = Order.objects.count()
    revenue = Order.objects.filter(
        payment_status='paid'
    ).aggregate(total=Sum('items__price'))['total'] or 0
    active_products = Product.objects.filter(is_available=True).count()
    low_stock = Inventory.objects.filter(
        quantity_on_hand__lte=F('reorder_points')
    ).count()

    # Orders last 30 days
    today = timezone.now().date()
    days = [(today - datetime.timedelta(days=i)) for i in range(29, -1, -1)]
    order_counts = [
        Order.objects.filter(created_at__date=day).count()
        for day in days
    ]

    # Recent orders
    recent_orders_data = []
    for order in Order.objects.select_related('user').prefetch_related(
        'items__product'
    ).order_by('-created_at')[:5]:
        first_item = order.items.first()
        recent_orders_data.append({
            'id': order.id,
            'customer': order.email,
            'product': first_item.product.name if first_item else '—',
            'row': order.address[:20] if order.address else '—',
            'qty': order.item_count,
            'sum': f'₱{order.total_price}',
            'date': order.created_at.strftime('%b %d, %Y'),
        })

    # Low stock items
    low_stock_items = list(
        Inventory.objects.filter(
            quantity_on_hand__lte=F('reorder_points')
        ).values('name', 'quantity_on_hand', 'unit', 'reorder_points')[:5]
    )
    # Out of stock items
    out_of_stock = Inventory.objects.filter(quantity_on_hand=0).count()

    # Top customers
    top_customers_data = [
        {'email': lp.user.email, 'points': lp.points}
        for lp in LoyaltyPoint.objects.select_related('user').order_by('-points')[:5]
    ]
    unread_notifications = Notification.objects.filter(is_read=False).order_by('-created_at')[:5]

    recent_logs = ActivityLog.objects.select_related('user').order_by('-created_at')[:8]
    context.update({
        'unread_count': Notification.objects.filter(is_read=False).count(),
        'unread_notifications': [
            {
                'id':         n.id, 
                'title':      n.title,
                'message':    n.message,
                'type':       n.type,
                'created_at': n.created_at.strftime('%b %d, %Y %H:%M'),
                'order_id':   n.order.id if n.order else None,
            }
            for n in unread_notifications
        ],
        "kpi": [
            {"title": "Total Orders", "metric": str(total_orders), "icon": "shopping_cart"},
            {"title": "Revenue", "metric": f"₱{revenue:,.2f}", "icon": "payments"},
            {"title": "Active Products", "metric": str(active_products), "icon": "coffee"},
            {"title": "Low Stock", "metric": str(low_stock) if low_stock else "—", "icon": "warning"},
        ],
        "chart": {
            "labels": [d.strftime('%b %d') for d in days],
            "datasets": [{
                "label": "Orders",
                "data": order_counts,
                "backgroundColor": "rgba(111, 78, 55, 0.7)",
                "borderColor": "#6f4e37",
                "borderWidth": 1,
            }],
        
        },
        'recent_logs': [
            {
                'action':     log.get_action_display(),
                'action_key': log.action,
                'user':       log.user.email if log.user else 'System',
                'details':    log.details,
                'created_at': log.created_at.strftime('%b %d, %H:%M'),
            }
            for log in recent_logs
        ],
        "recent_orders": recent_orders_data,
        'low_stock_items': low_stock_items,
        'out_of_stock':    out_of_stock,
        "top_customers": top_customers_data,
    })

    return context